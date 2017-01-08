import openpyxl

MAX_ROW = 200
MAX_COLUMN = 5

def get_total_income_per_month(sheet):
    MAX_ROW = sheet.max_row
    MAX_COLUMN = sheet.max_column
 
    total = 0
    for i in range(2, MAX_ROW):
        if(sheet.cell(row = i, column = 4).value is not None and sheet.cell(row = i, column = 2).value is not None):
            value = sheet.cell(row = i, column = 4).value
            total = total + float(value)

    return total

def  get_total_expense_per_month(sheet):
    MAX_ROW = sheet.max_row
    MAX_COLUMN = sheet.max_column
       
    total = 0
    for i in range(2, MAX_ROW):
        if(sheet.cell(row = i, column = 5).value is not None and sheet.cell(row = i, column = 2).value is not None):
            value = sheet.cell(row = i, column = 5).value
            total = total + float(value)
    return total

def get_totals_for_year(wb, year):
    total_income = 0
    total_expense = 0

    for sheet in wb.worksheets:
        if(year in sheet.title):
            total_income = total_income + get_total_income_per_month(sheet)
            total_expense = total_expense + get_total_expense_per_month(sheet)
    balance(total_income, total_expense)

def balance(income, expense):
    print "The total income is " + str(income)
    print "The total expense is " + str(expense)
    print "Remaining balance is " + str(income - expense)

def load_excel_workbook():
    wb = openpyxl.load_workbook('Expenses.xlsx', data_only = True)
    return wb

def display_menu_options(wb):
    menu = {}
    menu['1'] = "Display monthly statement"
    menu['2'] = "Display yearly statement"
    menu['3'] = "Get expense by category"

    while True:
        options = menu.keys()
        options.sort()
        for entry in options:
            print entry, menu[entry]
        
        selection = raw_input("Please select : ")
        if(selection == '1'):
            month = raw_input("Please enter month followed by year (October 2016) : ")
            sheet = wb.get_sheet_by_name(month)
            income_total = get_total_income_per_month(sheet)
            expense_total = get_total_expense_per_month(sheet)
            balance(income_total, expense_total)
        
        elif(selection == '2'):
            year = raw_input("Please enter year : ")
            get_totals_for_year(wb, year)
                    
        else:
            print "Unknown option selected"

def run():
    wb = load_excel_workbook()
    sheet = wb.get_sheet_by_name('October 2016')
    display_menu_options(wb)

if __name__ == '__main__':
    run()
